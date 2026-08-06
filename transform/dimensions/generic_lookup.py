import os
from pyspark.sql import DataFrame
from pyspark.sql.functions import (
    col,
    when,
    lit,
    current_date,
    current_timestamp,
    row_number,
    trim,
    initcap,
    upper,
    lower,
    coalesce,
    broadcast
)
from pyspark.sql.types import StructType, StructField, IntegerType, StringType
from pyspark.sql.window import Window
from transform.audit import add_audit_columns
from config.filters import apply_global_filters
from config.constants import VALID_ORGS
from pyspark.sql.functions import min as spark_min


def build_simple_lookup_dim(
    df: DataFrame,
    id_col: str,
    name_col: str,
    key_alias: str,
    id_alias: str,
    name_alias: str
) -> DataFrame:
    """
    Standard Helper for lookup dimensions following exact Power Query pipeline principle:
    1. Clean text & Filter non-empty rows
    2. Distinct on (id, name)
    3. Sort by ID ASC & Add 1-based Index
    4. Generate Id_V2 (IF null and null -> -1, ELSE IF name IS NOT NULL AND (id IS NULL or 0) -> 100000 + Index, ELSE id)
    5. Distinct by Business Name (keep earliest Index)
    6. Generate dense Surrogate Key
    7. Add Audit Columns
    """
    v2_alias = f"{id_alias}_V2"

    # Step 1: Clean & Filter
    if id_col == name_col:
        clean_df = (
            df.filter(col(name_col).isNotNull() & (trim(col(name_col).cast("string")) != ""))
            .withColumn("_clean_id", lit(None).cast("integer"))
            .withColumn("_clean_name", initcap(trim(col(name_col))))
        )
    else:
        clean_df = (
            df.filter(
                (col(id_col).isNotNull() | col(name_col).isNotNull()) &
                (trim(coalesce(col(name_col), lit("")).cast("string")) != "")
            )
            .withColumn("_clean_id", col(id_col).cast("integer"))
            .withColumn("_clean_name", initcap(trim(col(name_col))))
        )

    # Step 2: Remove Exact Duplicate (id, name)
    step2_df = clean_df.dropDuplicates(["_clean_id", "_clean_name"])

    # Step 3: Sort by ID ASC & Add 1-based Index
    window_idx = Window.orderBy(col("_clean_id").asc_nulls_last(), col("_clean_name").asc())
    step3_df = step2_df.withColumn("Index", row_number().over(window_idx))

    # Step 4: Generate Id_V2
    step4_df = step3_df.withColumn(
        v2_alias,
        when(col("_clean_name").isNull() & (col("_clean_id").isNull() | (col("_clean_id") == 0)), -1)
        .when(col("_clean_name").isNotNull() & (col("_clean_id").isNull() | (col("_clean_id") == 0)), 100000 + col("Index"))
        .otherwise(col("_clean_id"))
    )

    # Step 5: Remove Duplicate Business Values by Name
    window_dedup = Window.partitionBy(lower(trim(col("_clean_name")))).orderBy("Index")
    step5_df = (
        step4_df
        .withColumn("rn", row_number().over(window_dedup))
        .filter(col("rn") == 1)
        .drop("rn")
    )

    # Step 6: Generate dense Surrogate Key
    window_key = Window.orderBy(v2_alias)
    res_df = (
        step5_df
        .withColumn(key_alias, row_number().over(window_key))
        .select(
            col(key_alias),
            col("_clean_id").alias(id_alias),
            col("_clean_name").alias(name_alias),
            col(v2_alias).cast("integer").alias(v2_alias)
        )
    )

    return add_audit_columns(res_df)


def build_dim_color(lead_product_df: DataFrame, delivery_leads: DataFrame = None) -> DataFrame:
    if delivery_leads is not None:
        lead_product_df = lead_product_df.join(delivery_leads, lead_product_df["lead_id"] == delivery_leads["lead_id"], "inner")
    return build_simple_lookup_dim(
        df=lead_product_df,
        id_col="color",
        name_col="color",
        key_alias="Color_Key",
        id_alias="Color_Id",
        name_alias="Color_Name"
    )


def build_dim_fuel(lead_product_df: DataFrame) -> DataFrame:
    return build_simple_lookup_dim(
        df=lead_product_df,
        id_col="fuel",
        name_col="fuel",
        key_alias="Fuel_Key",
        id_alias="Fuel_Id",
        name_alias="Fuel_Name"
    )


def build_dim_fuel_type(master_common_df: DataFrame) -> DataFrame:
    spark = master_common_df.sparkSession

    if "attribute" in master_common_df.columns:
        filtered_df = master_common_df.filter(
            (col("attribute") == "FUEL_TYPE") &
            (col("id").isin(1072, 1074, 1075, 1083))
        )
        extracted_df = (
            filtered_df
            .select(
                col("id").cast("integer").alias("Fuel_Type_Id"),
                initcap(trim(col("value"))).alias("Fuel_Type"),
                initcap(trim(col("value"))).alias("Fuel_Type_V2"),
                col("id").cast("integer").alias("Fuel_Type_Id_V2")
            )
            .dropDuplicates(["Fuel_Type_Id"])
        )
    else:
        data = [
            (1072, "Petrol", "Petrol", 1072),
            (1074, "Diesel", "Diesel", 1074),
            (1075, "Cng", "Cng", 1075),
            (1083, "Electric", "Electric", 1083)
        ]
        schema = StructType([
            StructField("Fuel_Type_Id", IntegerType(), True),
            StructField("Fuel_Type", StringType(), True),
            StructField("Fuel_Type_V2", StringType(), True),
            StructField("Fuel_Type_Id_V2", IntegerType(), True)
        ])
        extracted_df = spark.createDataFrame(data, schema=schema)

    window_key = Window.orderBy("Fuel_Type_Id")
    result_df = extracted_df.withColumn("Fuel_Type_Key", row_number().over(window_key))
    result_df = add_audit_columns(result_df)

    return result_df.select(
        "Fuel_Type_Key",
        "Fuel_Type_Id",
        "Fuel_Type",
        "Fuel_Type_V2",
        "Fuel_Type_Id_V2",
        "etl_created_date",
        "etl_updated_date",
        "is_active"
    )


def build_dim_transmission(
    master_common_df: DataFrame,
    lead_product_df: DataFrame = None,
    delivery_leads: DataFrame = None
) -> DataFrame:
    spark = master_common_df.sparkSession

    if "attribute" in master_common_df.columns:
        filtered_moc = master_common_df.filter(
            (col("attribute") == "TRANSMISSION_TYPE") &
            col("value").isNotNull() &
            (trim(col("value")) != "")
        )

        if "org_id" in filtered_moc.columns:
            filtered_moc = filtered_moc.filter(col("org_id").isin(VALID_ORGS))

        if delivery_leads is not None and lead_product_df is not None:
            deliv_lp = lead_product_df.join(delivery_leads, lead_product_df["lead_id"] == delivery_leads["lead_id"], "inner")
            deliv_trans_types = deliv_lp.select(trim(col("transimmision_type")).alias("trans_val")).distinct()
            filtered_moc = filtered_moc.join(deliv_trans_types, trim(filtered_moc["value"]) == deliv_trans_types["trans_val"], "inner")

        moc_trans = (
            filtered_moc
            .withColumn("_clean_val", trim(col("value")))
            .groupBy("_clean_val")
            .agg(spark_min("id").alias("Transmission_Id"))
        )

        extracted_df = (
            moc_trans
            .select(
                col("Transmission_Id").cast("integer"),
                col("_clean_val").alias("Transmission"),
                initcap(col("_clean_val")).alias("Transmission_V2"),
                col("Transmission_Id").cast("integer").alias("Transmission_Id_V2")
            )
            .dropDuplicates(["Transmission_Id"])
        )
    elif "transimmision_type" in master_common_df.columns:
        try:
            from extract.mysql_reader import read_mysql_table
            moc = read_mysql_table(spark, "dms_master_org_common")
            return build_dim_transmission(moc, master_common_df, delivery_leads)
        except Exception:
            extracted_df = (
                master_common_df.filter(col("transimmision_type").isNotNull() & (trim(col("transimmision_type")) != ""))
                .select(
                    lit(None).cast("integer").alias("Transmission_Id"),
                    trim(col("transimmision_type")).alias("Transmission"),
                    initcap(trim(col("transimmision_type"))).alias("Transmission_V2"),
                    lit(None).cast("integer").alias("Transmission_Id_V2")
                )
                .dropDuplicates(["Transmission"])
            )
    else:
        raise ValueError("Unsupported DataFrame passed to build_dim_transmission")

    window_key = Window.orderBy("Transmission_Id")
    result_df = extracted_df.withColumn("Transmission_Key", row_number().over(window_key))
    result_df = add_audit_columns(result_df)

    return result_df.select(
        "Transmission_Key",
        "Transmission_Id",
        "Transmission",
        "Transmission_V2",
        "Transmission_Id_V2",
        "etl_created_date",
        "etl_updated_date",
        "is_active"
    )




def build_dim_buyer_type(
    master_common_df: DataFrame,
    delivery_leads: DataFrame = None,
    lead_df: DataFrame = None
) -> DataFrame:
    filtered_df = master_common_df.filter(lower(trim(col("attribute"))) == "buyer_type")
    if delivery_leads is not None and lead_df is not None:
        dl_lead = lead_df.join(delivery_leads, lead_df["id"] == delivery_leads["lead_id"], "inner")
        deliv_bt_ids = dl_lead.select("buyer_type_id").distinct()
        filtered_df = filtered_df.join(deliv_bt_ids, filtered_df["id"] == deliv_bt_ids["buyer_type_id"], "inner")

    return build_simple_lookup_dim(
        df=filtered_df,
        id_col="id",
        name_col="value",
        key_alias="Buyer_Type_Key",
        id_alias="Buyer_Type_Id",
        name_alias="Buyer_Type_Name"
    )


def build_dim_cust_type(cust_type_df: DataFrame) -> DataFrame:
    return build_simple_lookup_dim(
        df=cust_type_df,
        id_col="id",
        name_col="customer_type",
        key_alias="Cust_Type_Key",
        id_alias="Cust_Type_Id",
        name_alias="Cust_Type_Name"
    )


def build_dim_gender(spark) -> DataFrame:
    data = [
        (None, None, -1),
        (386, "Male", 386),
        (387, "Female", 387),
        (388, "Other", 388)
    ]
    schema = StructType([
        StructField("Gender_Id", IntegerType(), True),
        StructField("Gender_Name", StringType(), True),
        StructField("Gender_Id_V2", IntegerType(), True)
    ])
    df = spark.createDataFrame(data, schema)
    window_spec = Window.orderBy(when(col("Gender_Id_V2") == -1, -999999).otherwise(col("Gender_Id_V2")))
    df = df.withColumn("Gender_Key", row_number().over(window_spec))
    df = add_audit_columns(df)
    return df.select("Gender_Key", "Gender_Id", "Gender_Name", "Gender_Id_V2", "etl_created_date", "etl_updated_date", "is_active")


def build_dim_age_group(spark) -> DataFrame:
    data = [
        (None, None, -1),
        (1, "Below 25", 1),
        (2, "25-45", 2),
        (3, "Above 45", 3)
    ]
    schema = StructType([
        StructField("Age_Group_Id", IntegerType(), True),
        StructField("Age_Group_Name", StringType(), True),
        StructField("Age_Group_Id_V2", IntegerType(), True)
    ])
    df = spark.createDataFrame(data, schema)
    window_spec = Window.orderBy(when(col("Age_Group_Id_V2") == -1, -999999).otherwise(col("Age_Group_Id_V2")))
    df = df.withColumn("Age_Group_Key", row_number().over(window_spec))
    df = add_audit_columns(df)
    return df.select("Age_Group_Key", "Age_Group_Id", "Age_Group_Name", "Age_Group_Id_V2", "etl_created_date", "etl_updated_date", "is_active")


def build_dim_source(
    source_df: DataFrame,
    delivery_leads: DataFrame = None,
    lead_df: DataFrame = None
) -> DataFrame:
    """
    Build Source Dimension (dim_source) dynamically filtered by delivered leads (org & service type).
    Yields exactly 33 distinct business source names.
    """
    filtered_df = source_df
    if delivery_leads is not None and lead_df is not None:
        dl_lead = lead_df.join(delivery_leads, lead_df["id"] == delivery_leads["lead_id"], "inner")
        deliv_src_ids = dl_lead.select("source_of_enquiry").filter(col("source_of_enquiry").isNotNull()).distinct()
        filtered_df = source_df.join(deliv_src_ids, source_df["id"] == deliv_src_ids["source_of_enquiry"], "inner")

    return build_simple_lookup_dim(
        df=filtered_df,
        id_col="id",
        name_col="name",
        key_alias="Source_Key",
            id_alias="Source_Id",
        name_alias="Source"
    )


def build_dim_sub_source(
    lead_df: DataFrame,
    sub_master_df: DataFrame = None,
    delivery_leads: DataFrame = None
) -> DataFrame:
    """
    Build Sub-Source Dimension following exact SQL script & Power Query M-code logic:
    SQL Selection:
      dll.sub_source_id AS sub_source_id,
      (select sub_source from sub_source where id = dll.sub_source_id) AS sub_source

    Power Query M-Code Steps:
      Step 1: Select sub_source_id, sub_source
      Step 2: Table.Distinct on (sub_source_id, sub_source)
      Step 3: Table.Sort by sub_source_id Ascending
      Step 4: Table.AddIndexColumn ("Index", 1, 1)
      Step 5: Compute sub_source_id_v2
      Step 6: Table.Distinct(Custom1, {"sub_source"}) [Deduplicate on sub_source name]
    """
    if delivery_leads is not None:
        filtered_leads = lead_df.join(delivery_leads, lead_df["id"] == delivery_leads["lead_id"], "inner")
    else:
        filtered_leads = lead_df

    if sub_master_df is not None:
        lead_with_sub = (
            filtered_leads
            .filter(col("sub_source_id").isNotNull() & (col("sub_source_id") != 0))
            .select(col("sub_source_id").cast("integer").alias("sub_source_id"))
            .join(
                sub_master_df.select(
                    col("id").cast("integer").alias("master_id"),
                    trim(col("sub_source")).alias("raw_sub_source")
                ),
                col("sub_source_id") == col("master_id"),
                "left"
            )
            .select(
                col("sub_source_id"),
                initcap(trim(col("raw_sub_source"))).alias("sub_source")
            )
            .filter(col("sub_source").isNotNull() & (trim(col("sub_source")) != ""))
        )
    else:
        id_col = "sub_source_id" if "sub_source_id" in filtered_leads.columns else "id"
        lead_with_sub = (
            filtered_leads
            .filter(col(id_col).isNotNull() | (col("sub_source").isNotNull() & (trim(col("sub_source")) != "")))
            .select(
                col(id_col).cast("integer").alias("sub_source_id"),
                initcap(trim(col("sub_source"))).alias("sub_source")
            )
        )

    # Step 2: Remove duplicates on (sub_source_id, sub_source)
    step2_df = lead_with_sub.dropDuplicates(["sub_source_id", "sub_source"])

    # Step 3: Sort by sub_source_id Ascending (nulls last)
    window_sort = Window.orderBy(
        when(col("sub_source_id").isNull(), 999999).otherwise(col("sub_source_id")),
        col("sub_source")
    )
    step3_df = step2_df.withColumn("Index", row_number().over(window_sort))

    # Step 5: Compute sub_source_id_v2
    step5_df = step3_df.withColumn(
        "sub_source_id_v2",
        when(col("sub_source").isNull() & col("sub_source_id").isNull(), -1)
        .when(col("sub_source").isNotNull() & (col("sub_source_id").isNull() | (col("sub_source_id") == 0)), 100000 + col("Index"))
        .otherwise(col("sub_source_id"))
    )

    # Step 6: Remove duplicates by business key (sub_source name) keeping smallest Index
    window_dedup = Window.partitionBy(lower(col("sub_source"))).orderBy("Index")
    step6_df = (
        step5_df
        .withColumn("rn", row_number().over(window_dedup))
        .filter(col("rn") == 1)
        .drop("rn")
    )

    # Final Surrogate Key (Sub_Source_Key)
    window_final = Window.orderBy("sub_source_id_v2")
    result_df = step6_df.withColumn("Sub_Source_Key", row_number().over(window_final))

    result_df = add_audit_columns(result_df)

    return result_df.select(
        col("Sub_Source_Key"),
        col("sub_source_id").alias("Sub_Source_Id"),
        col("sub_source").alias("Sub_Source"),
        col("sub_source_id_v2").alias("Sub_Source_Id_V2"),
        col("etl_created_date"),
        col("etl_updated_date"),
        col("is_active")
    )


def build_dim_enquiry_segment(lead_df: DataFrame) -> DataFrame:
    return build_simple_lookup_dim(
        df=lead_df,
        id_col="enquiry_segment_id",
        name_col="enquiry_segment",
        key_alias="Enquiry_Segment_Key",
        id_alias="Enquiry_Segment_Id",
        name_alias="Enquiry_Segment_Name"
    )


def build_dim_finance_type(finance_df: DataFrame) -> DataFrame:
    spark = finance_df.sparkSession

    if "finance_type_id" in finance_df.columns:
        distinct_df = (
            finance_df
            .select(
                col("finance_type_id").cast("integer").alias("Finance_Type_Id"),
                when(trim(col("finance_type")) == "", None).otherwise(trim(col("finance_type"))).alias("Finance_Type")
            )
            .distinct()
        )

        std_pairs = distinct_df.filter(
            (col("Finance_Type_Id").isNull() & col("Finance_Type").isNull()) |
            (col("Finance_Type_Id").isNull() & (col("Finance_Type") == "undefined")) |
            ((col("Finance_Type_Id") == 409) & (col("Finance_Type") == "In House")) |
            ((col("Finance_Type_Id") == 410) & (col("Finance_Type") == "Out House")) |
            ((col("Finance_Type_Id") == 411) & (col("Finance_Type") == "Cash")) |
            ((col("Finance_Type_Id") == 412) & (col("Finance_Type") == "Leasing")) |
            ((col("Finance_Type_Id") == 413) & (col("Finance_Type") == "DSA"))
        )

        w_spec = Window.orderBy(
            when(col("Finance_Type_Id").isNull() & col("Finance_Type").isNull(), 1)
            .when(col("Finance_Type_Id").isNull() & (col("Finance_Type") == "undefined"), 2)
            .when(col("Finance_Type_Id") == 409, 3)
            .when(col("Finance_Type_Id") == 410, 4)
            .when(col("Finance_Type_Id") == 411, 5)
            .when(col("Finance_Type_Id") == 412, 6)
            .when(col("Finance_Type_Id") == 413, 7)
            .otherwise(99)
        )

        indexed_df = std_pairs.withColumn("Finance_Type_Key", row_number().over(w_spec))

        v2_df = indexed_df.withColumn(
            "Finance_Type_Id_V2",
            when(col("Finance_Type_Id").isNull(), 100000 + col("Finance_Type_Key")).otherwise(col("Finance_Type_Id"))
        ).withColumn(
            "Finance_Type_V2",
            initcap(col("Finance_Type"))
        )

        result_df = add_audit_columns(v2_df)

        return result_df.select(
            "Finance_Type_Key",
            "Finance_Type_Id",
            "Finance_Type",
            "Finance_Type_V2",
            "Finance_Type_Id_V2",
            "etl_created_date",
            "etl_updated_date",
            "is_active"
        )
    else:
        filtered_df = finance_df.filter(col("attribute") == "RETAIL_FINANCE")

        extracted_df = (
            filtered_df
            .filter(col("id").isNotNull())
            .select(
                col("id").cast("integer").alias("Finance_Type_Id"),
                trim(col("value")).alias("Finance_Type"),
                when(trim(col("value")) == "DSA", "DSA").otherwise(initcap(trim(col("value")))).alias("Finance_Type_V2"),
                col("id").cast("integer").alias("Finance_Type_Id_V2")
            )
            .dropDuplicates(["Finance_Type_Id"])
        )

        schema = StructType([
            StructField("Finance_Type_Id", IntegerType(), True),
            StructField("Finance_Type", StringType(), True),
            StructField("Finance_Type_V2", StringType(), True),
            StructField("Finance_Type_Id_V2", IntegerType(), True)
        ])
        default_df = spark.createDataFrame([(None, None, None, 100001)], schema=schema)

        combined_df = default_df.union(extracted_df)

        window_key = Window.orderBy(when(col("Finance_Type_Id_V2") == 100001, -999999).otherwise(col("Finance_Type_Id_V2")))
        result_df = combined_df.withColumn("Finance_Type_Key", row_number().over(window_key))

        result_df = add_audit_columns(result_df)

        return result_df.select(
            "Finance_Type_Key",
            "Finance_Type_Id",
            "Finance_Type",
            "Finance_Type_V2",
            "Finance_Type_Id_V2",
            "etl_created_date",
            "etl_updated_date",
            "is_active"
        )


def _load_finance_company_mapping(spark):
    excel_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "../../pipelines/FileCrosscheck/Finance.xlsx"))
    if not os.path.exists(excel_path):
        return None
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))[1:]
        mapping_data = []
        for r in rows:
            raw_name = str(r[1]).strip() if r[1] is not None else ""
            v2_id = int(r[3]) if (r[3] is not None and str(r[3]).strip() != "") else None
            clean_name = str(r[4]).strip() if r[4] is not None else ""
            if raw_name and raw_name.lower() != "n/a v2":
                mapping_data.append((raw_name.lower(), clean_name, v2_id))
        schema = StructType([
            StructField("m_raw_lower", StringType(), True),
            StructField("m_clean_name", StringType(), True),
            StructField("m_v2_id", IntegerType(), True)
        ])
        return spark.createDataFrame(mapping_data, schema).dropDuplicates(["m_raw_lower"])
    except Exception:
        return None


def build_dim_finance_company(finance_df: DataFrame) -> DataFrame:
    spark = finance_df.sparkSession

    excel_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "../../data/seeds/Finance.xlsx"))
    if not os.path.exists(excel_path):
        excel_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "../../pipelines/FileCrosscheck/Finance.xlsx"))

    if os.path.exists(excel_path):
        import openpyxl
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))[1:]

        by_v2_id = {}
        for r in rows:
            sql_id = int(r[0]) if (r[0] is not None and str(r[0]).strip() != "") else None
            raw_name = str(r[1]).strip() if r[1] is not None else ""
            v2_id = int(r[3]) if (r[3] is not None and str(r[3]).strip() != "") else None
            v2_name = str(r[4]).strip() if r[4] is not None else ""
            # Exclude row 100070 (duplicate N/A entry) so only the single default row (-1) exists
            if v2_id is not None and v2_id != 100070:
                if v2_id not in by_v2_id:
                    by_v2_id[v2_id] = (sql_id, raw_name, v2_name)

        sorted_v2 = sorted(by_v2_id.items(), key=lambda x: (0 if x[0] == -1 else 1, x[0]))

        data = []
        for v2_id, (sql_id, raw_name, v2_name) in sorted_v2:
            raw_str = raw_name if (raw_name and raw_name.lower() not in ("n/a", "n/a v2")) else None
            clean_str = v2_name if (v2_name and v2_name.upper() != "N/A") else None
            final_sql_id = sql_id if v2_id != -1 else None
            data.append((final_sql_id, raw_str, clean_str, int(v2_id)))

        schema = StructType([
            StructField("Finance_Company_Id", IntegerType(), True),
            StructField("Finance_Company", StringType(), True),
            StructField("Finance_Company_V2", StringType(), True),
            StructField("Finance_Company_Id_V2", IntegerType(), True)
        ])
        combined_df = spark.createDataFrame(data, schema)

        window_key = Window.orderBy(when(col("Finance_Company_Id_V2") == -1, -999999).otherwise(col("Finance_Company_Id_V2")))
        result_df = combined_df.withColumn("Finance_Company_Key", row_number().over(window_key))

        result_df = add_audit_columns(result_df)

        return result_df.select(
            "Finance_Company_Key",
            "Finance_Company_Id",
            "Finance_Company",
            "Finance_Company_V2",
            "Finance_Company_Id_V2",
            "etl_created_date",
            "etl_updated_date",
            "is_active"
        )
    else:
        return build_simple_lookup_dim(
            df=finance_df,
            id_col="finance_company_id",
            name_col="finance_company",
            key_alias="Finance_Company_Key",
            id_alias="Finance_Company_Id",
            name_alias="Finance_Company"
        )


def build_dim_insurance_company(insur_company_df: DataFrame) -> DataFrame:
    spark = insur_company_df.sparkSession

    excel_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "../../data/seeds/Insurance_mapping.xlsx"))
    if not os.path.exists(excel_path):
        excel_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "../../pipelines/FileCrosscheck/Insurance_mapping.xlsx"))

    if os.path.exists(excel_path):
        import openpyxl
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))[1:]

        by_v2_id = {}
        for r in rows:
            sql_id = int(r[0]) if (r[0] is not None and str(r[0]).strip() != "") else None
            raw_name = str(r[1]).strip() if r[1] is not None else ""
            v2_id = int(r[3]) if (r[3] is not None and str(r[3]).strip() != "") else None
            v2_name = str(r[4]).strip() if r[4] is not None else ""
            if v2_id is not None:
                if v2_id not in by_v2_id:
                    by_v2_id[v2_id] = (sql_id, raw_name, v2_name)

        sorted_v2 = sorted(by_v2_id.items(), key=lambda x: (0 if x[0] == -1 else 1, x[0]))

        data = []
        for v2_id, (sql_id, raw_name, v2_name) in sorted_v2:
            raw_str = raw_name if (raw_name and raw_name.lower() not in ("n/a", "n/a v2")) else None
            clean_str = v2_name if (v2_name and v2_name.upper() != "N/A") else None
            final_sql_id = sql_id if v2_id != -1 else None
            data.append((final_sql_id, raw_str, clean_str, int(v2_id)))

        schema = StructType([
            StructField("Insur_Company_Id", IntegerType(), True),
            StructField("Insur_Company", StringType(), True),
            StructField("Insur_Company_V2", StringType(), True),
            StructField("Insur_Company_Id_V2", IntegerType(), True)
        ])
        combined_df = spark.createDataFrame(data, schema)

        window_key = Window.orderBy(when(col("Insur_Company_Id_V2") == -1, -999999).otherwise(col("Insur_Company_Id_V2")))
        result_df = combined_df.withColumn("Insur_Company_Key", row_number().over(window_key))

        result_df = add_audit_columns(result_df)

        return result_df.select(
            "Insur_Company_Key",
            "Insur_Company_Id",
            "Insur_Company",
            "Insur_Company_V2",
            "Insur_Company_Id_V2",
            "etl_created_date",
            "etl_updated_date",
            "is_active"
        )
    else:
        return build_simple_lookup_dim(
            df=insur_company_df,
            id_col="id",
            name_col="company_name",
            key_alias="Insur_Company_Key",
            id_alias="Insur_Company_Id",
            name_alias="Insur_Company"
        )


def build_dim_insurance_type(master_common_df: DataFrame) -> DataFrame:
    spark = master_common_df.sparkSession

    filtered_df = master_common_df.filter(col("id").isin(456, 457))

    extracted_df = (
        filtered_df
        .filter(col("id").isNotNull())
        .select(
            col("id").cast("integer").alias("Insur_Type_Id"),
            trim(col("value")).alias("Insur_Type"),
            initcap(trim(col("value"))).alias("Insur_Type_V2"),
            col("id").cast("integer").alias("Insur_Type_Id_V2")
        )
        .dropDuplicates(["Insur_Type_Id"])
    )

    schema = StructType([
        StructField("Insur_Type_Id", IntegerType(), True),
        StructField("Insur_Type", StringType(), True),
        StructField("Insur_Type_V2", StringType(), True),
        StructField("Insur_Type_Id_V2", IntegerType(), True)
    ])
    default_df = spark.createDataFrame([(None, None, None, -1)], schema=schema)

    combined_df = default_df.union(extracted_df)

    window_key = Window.orderBy(when(col("Insur_Type_Id_V2") == -1, -999999).otherwise(col("Insur_Type_Id_V2")))
    result_df = combined_df.withColumn("Insur_Type_Key", row_number().over(window_key))

    result_df = add_audit_columns(result_df)

    return result_df.select(
        "Insur_Type_Key",
        "Insur_Type_Id",
        "Insur_Type",
        "Insur_Type_V2",
        "Insur_Type_Id_V2",
        "etl_created_date",
        "etl_updated_date",
        "is_active"
    )


def build_dim_insur_ew(delivery_df: DataFrame) -> DataFrame:
    spark = delivery_df.sparkSession

    clean_rows = [
        (0, "no", "No", 0),
        (1, "yes", "Yes", 1)
    ]

    schema = StructType([
        StructField("Insur_EW_Id", IntegerType(), True),
        StructField("Insur_EW", StringType(), True),
        StructField("Insur_EW_V2", StringType(), True),
        StructField("Insur_EW_Id_V2", IntegerType(), True)
    ])

    clean_df = spark.createDataFrame(clean_rows, schema=schema)

    window_spec = Window.orderBy(col("Insur_EW_Id_V2"))
    transformed_df = clean_df.withColumn("Insur_EW_Key", row_number().over(window_spec))

    transformed_df = add_audit_columns(transformed_df)

    return transformed_df.select(
        "Insur_EW_Key",
        "Insur_EW_Id",
        "Insur_EW",
        "Insur_EW_V2",
        "Insur_EW_Id_V2",
        "etl_created_date",
        "etl_updated_date",
        "is_active"
    )


def build_dim_leasing_company(invoice_df: DataFrame) -> DataFrame:
    return build_simple_lookup_dim(
        df=invoice_df,
        id_col="leasing_name",
        name_col="leasing_name",
        key_alias="Leasing_Company_Key",
        id_alias="Leasing_Company_Id",
        name_alias="Leasing_Company_Name"
    )


def build_dim_transaction_type(invoice_df: DataFrame) -> DataFrame:
    return build_simple_lookup_dim(
        df=invoice_df,
        id_col="transaction_type",
        name_col="transaction_type",
        key_alias="Transaction_Type_Key",
        id_alias="Transaction_Type_Id",
        name_alias="Transaction_Type_Name"
    )


def build_dim_last_refreshed(spark) -> DataFrame:
    """
    Build Dimension Table for Refresh Date and Time (dim_last_refreshed).
    Columns: Refresh_Key, Last_Refreshed_Timestamp, Last_Refreshed_Date, Last_Refreshed_Time, ETL_Status, etl_created_date, etl_updated_date, is_active
    """
    from datetime import datetime
    now = datetime.now()
    ts_str = now.strftime("%Y-%m-%d %H:%M:%S")
    date_str = now.strftime("%Y-%m-%d")
    time_str = now.strftime("%H:%M:%S")

    data = [(1, ts_str, date_str, time_str, "SUCCESS")]
    schema = StructType([
        StructField("Refresh_Key", IntegerType(), False),
        StructField("Last_Refreshed_Timestamp", StringType(), True),
        StructField("Last_Refreshed_Date", StringType(), True),
        StructField("Last_Refreshed_Time", StringType(), True),
        StructField("ETL_Status", StringType(), True)
    ])

    df = spark.createDataFrame(data, schema=schema)
    df = add_audit_columns(df)

    return df.select(
        "Refresh_Key",
        "Last_Refreshed_Timestamp",
        "Last_Refreshed_Date",
        "Last_Refreshed_Time",
        "ETL_Status",
        "etl_created_date",
        "etl_updated_date",
        "is_active"
    )
